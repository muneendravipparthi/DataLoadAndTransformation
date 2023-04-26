import numpy as np
import pandas as pd
from openpyxl.utils.cell import get_column_letter

input_file_1 = "../app/in/InstrumentDetails.csv"
input_file_2 = "../app/in/PositionDetails.csv"
output_file = "../app/out/PositionReport.csv"


class Data_validation:

    def dv(self):
        print("*******Execution Started*******")
        # reading CSV
        InstrumentDetails_df = pd.read_csv(input_file_1)
        PositionDetails_df = pd.read_csv(input_file_2)
        # converting the inputs to output format for source and destination validation
        df = self.Generate_OutputFile(InstrumentDetails_df, PositionDetails_df)
        PositionReport_df = pd.read_csv(output_file)
        # comparing source and destination data
        df = self.Find_Differences(PositionReport_df, df)
        #R report generation with differences
        self.Report_Generation(df)
        print("*******Execution Completed*******")

    def Generate_OutputFile(self, InstrumentDetails_df, PositionDetails_df):
        PositionDetails_df = PositionDetails_df.rename(columns={'ID': 'PositionID'}, inplace=False)
        output_df = pd.merge(InstrumentDetails_df, PositionDetails_df, how='right', left_on='ISIN',
                             right_on='InstrumentID')
        output_df['Total Price'] = output_df['Quantity'] * output_df['Unit Price']
        output_df['ID'] = output_df.index + 1
        output_df["ID"] = output_df.ID.map("{:03}".format)
        output_df['ID'] = 'PR' + output_df['ID'].astype(str)
        outputcolumnslist = ['ID', 'PositionID', 'ISIN', 'Quantity', 'Total Price']
        output_df = output_df[outputcolumnslist]
        return output_df

    def Find_Differences(self, df1, df2):
        df1.equals(df2)
        comparison_values = df1.values == df2.values
        rows, cols = np.where(comparison_values == False)

        for item in zip(rows, cols):
            df1.iloc[item[0], item[1]] = '{} --> {}'.format(df1.iloc[item[0], item[1]], df2.iloc[item[0], item[1]])

        df1.to_csv('PositionReport_validationreport.csv', index=False)
        return df1

    def Report_Generation(self, df1):
        file_name = "PositionReport_validationreport.xlsx"
        sheet_name = "report"
        # Insert an empty column to write the formulas
        df1.insert(len(df1.columns), 'Execution_Status', np.nan)

        writer = pd.ExcelWriter(file_name, engine='xlsxwriter')

        # Convert the dataframe to an XlsxWriter Excel object.
        df1.to_excel(writer, sheet_name=sheet_name, index=False, header=True)

        # Get the xlsxwriter workbook and worksheet objects.
        workbook = writer.book
        worksheet = writer.sheets[sheet_name]
        worksheet.freeze_panes(1, 1)

        # Create a for loop to start writing the formulas to each row
        for row in range(2, df1.shape[0] + 2):
            formula = f'=IF(COUNTIF(A{row}: {get_column_letter(len(df1.columns) - 1)}{row}, "*-->*"), "FAIL", "PASS")'

            worksheet.write_formula(f"{get_column_letter(len(df1.columns))}{row}", formula)

        status_column_position = get_column_letter(len(df1.columns))

        # Add a format. Light red fill with dark red text.
        format1 = workbook.add_format({'bg_color': '#FFC7CE', 'font_color': '#9C0006'})
        format5 = workbook.add_format({'bg_color': '#FFDEAD', 'font_color': '#000000'})

        # Apply a conditional format to the cell range.
        worksheet.conditional_format(1, 0, len(df1), len(df1.columns),
                                     {'type': 'text',
                                      'criteria': 'containing',
                                      'value': '-->',
                                      'format': format1})

        worksheet.conditional_format(1, 0, len(df1), len(df1.columns),
                                     {'type': 'text',
                                      'criteria': 'containing',
                                      'value': '<ignore>',
                                      'format': format5})

        # Add a format. Blue fill with White text.
        format2 = workbook.add_format({'bg_color': '#1589FF', 'font_color': '#FFFFFF'})

        worksheet.conditional_format(0, 0, 0, len(df1.columns) - 1,
                                     {'type': 'text', 'criteria': 'not containing', 'value': '-->', 'format': format2})

        # Add a format. Green fill with White text.
        format3 = workbook.add_format({'bg_color': '#41A317', 'font_color': '#FFFFFF'})

        # Add a format. Red fill with White text.
        format4 = workbook.add_format({'bg_color': '#E41B17', 'font_color': '#FFFFFF'})

        worksheet.conditional_format(1, len(df1.columns) - 1, len(df1), len(df1.columns) - 1,
                                     {'type': 'text',
                                      'criteria': 'containing',
                                      'value': 'PASS',
                                      'format': format3})

        worksheet.conditional_format(1, len(df1.columns) - 1, len(df1), len(df1.columns) - 1,
                                     {'type': 'text',
                                      'criteria': 'containing',
                                      'value': 'FAIL',
                                      'format': format4})

        # Close the Pandas Excel writer and output the Excel file.
        writer._save()


datavalidation = Data_validation()
datavalidation.dv()
